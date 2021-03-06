import os
import re
import sys
import time
import random
import logging
import yaml
import requests
import json
import pandas as pd
import argparse

from datetime import date, datetime
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import colors, Font

def get_useragent(ua_pool):
    return random.choice(ua_pool)

def check_proxy(address, ip_check):
    proxies = {'http': address, 'https': address}
    try:
        requests.get(ip_check, proxies=proxies)
        return True
    except:
        return False

def get_proxies(proxy_list, ip_check, max_proxies=5):
    logging.info('Getting proxies...')
    proxies = []

    r = requests.get(proxy_list)
    soup = BeautifulSoup(r.content, 'lxml')

    ip_table = soup.find('tbody')
    for row in ip_table:
        row  = row.findAll('td')
        ip   = row[0].text
        port = row[1].text
        anon = row[4].text
        http = row[6].text

        address = '{}:{}'.format(ip, port)
        if anon != 'transparent' and http == 'yes' and check_proxy(address, ip_check):
            proxies.append(address)
            logging.info('Adding proxy {}/{}: {}'.format(len(proxies), max_proxies, address))
        if len(proxies) >= max_proxies:
            break

    return proxies

def scrape(products, headers, ua_pool, do_proxies=True, proxy_list=None, max_proxies=5, ip_check=None, verbose=False):
    rex = re.compile('\$(\d+.\d+)')

    data = {}
    error = []

    if do_proxies:
        proxy_pool = get_proxies(proxy_list, ip_check, max_proxies)

    for category in products:
        data[category['name']] = {}
        for i, asin in enumerate(category['asins']):
            logging.info('Getting info for product {}...'.format(asin))

            try:
                while True:
                    # set ip/headers
                    headers['user-agent'] = get_useragent(ua_pool)
                    if do_proxies:
                        address = random.choice(proxy_pool)
                        proxies = {'http': address, 'https': address}

                    # make request
                    if do_proxies:
                        try:
                            r = requests.get('https://amazon.com/dp/{}'.format(asin),
                                             headers=headers,
                                             proxies=proxies)
                        except:
                            logging.error('Proxy failing... using local IP instead')
                            r = requests.get('https://amazon.com/dp/{}'.format(asin),
                                             headers=headers)
                    else:
                        r = requests.get('https://amazon.com/dp/{}'.format(asin),
                                         headers=headers)

                    soup = BeautifulSoup(r.content, 'lxml')
                    access_date = date.today()

                    if verbose:
                        print(r.content)
                        print()

                    # check CAPTCHA detection
                    captcha = soup.find('title')
                    if captcha is not None and captcha.text == 'Robot Check':
                        logging.error('Robot detected! Please do the CAPTCHA in browser...')
                        input('Press any key to continue (press Ctrl-C to quit)...')
                    else:
                        break

                # search html
                title = soup.find('span', attrs={'id': 'productTitle'})
                prime = soup.find('a', attrs={'href': re.compile('^/gp/prime/pipeline/signup.html?')})
                prime_price = soup.find('span', attrs={'id': 'primeExclusivePricingMessage'})
                price = soup.find('span', attrs={'id': 'priceblock_dealprice'})
                shipping = soup.find('span', attrs={'id': 'ourprice_shippingmessage'})
                shipping_detail = soup.find('div', attrs={'id': 'a-popover-shippingDetailsDisplayContent'})
                merchant = soup.find('div', attrs={'id': 'merchant-info'})
                rating = soup.find('span', attrs={'id': 'acrPopover'})
                n_ratings = soup.find('span', attrs={'id': 'acrCustomerReviewText'})
                availability = soup.find('div', attrs={'id': 'availability'})
                # missing quantity + packaging info

                seller = None
                fulfiller = None

                # parse info
                if title is not None:
                    title = title.text.strip()
                if prime is not None:
                    prime = 'Yes'
                else:
                    prime = 'No'
                if price is not None:
                    price = price.text.strip()
                else:
                    price = soup.find('span', attrs={'id': 'priceblock_ourprice'})
                    if price is not None:
                        price = price.text.strip()
                if prime_price is not None and price is not None:
                    prime_discount = rex.search(prime_price.text)
                    if prime_discount is not None:
                        prime_discount = prime_discount.group()
                        prime_price = float(price)-float(prime_discount)
                elif prime_price is None:
                    prime_price = price
                if shipping is not None:
                    shipping = shipping.text.strip()
                    if 'FREE' in shipping:
                        shipping = '$0'
                    else:
                        shipping = rex.search(shipping)
                        if shipping is not None:
                            shipping = shipping.group()
                if shipping_detail is not None:
                    detail = shipping_detail.findAll('span')
                    if detail is not None and len(detail) == 3:
                        shipping = rex.search(detail[2].text)
                        if shipping is not None:
                            shipping = shipping.group()
                if merchant is not None:
                    merchant = merchant.text.strip()
                    if 'Ships from and sold by ' in merchant:
                        seller = merchant[len('Ships from and sold by '):-1]
                        fulfiller = seller
                    else:
                        ind = merchant.find(' and ')
                        seller = merchant[len('Sold by '):ind]
                        fulfiller = merchant[ind+len(' and Fulfilled by '):-1]
                if rating is not None:
                    rating = float(rating.attrs['title'].strip().split()[0])
                else:
                    rating = 'NA'
                if n_ratings is not None:
                    n_ratings = int(n_ratings.text.strip().split()[0].replace(',',''))
                else:
                    n_ratings = 0
                if availability is not None:
                    availability = availability.text.strip()
                    if availability == 'Currently Unavailable.':
                        seller = 'UNAVAILABLE'
                        fulfiller = 'UNAVAILABLE'

                # save data
                url = 'https://www.amazon.com/dp/{}'.format(asin)
                data[category['name']][i] = [access_date,
                                             ' ', # search term column
                                             title,
                                             '=HYPERLINK("{}","{}")'.format(url, url),
                                             ' ', # quantity column
                                             prime,
                                             prime_price,
                                             price,
                                             shipping,
                                             seller,
                                             fulfiller,
                                             rating,
                                             n_ratings,
                                             ' '] # packaging column

                if any(x is None for x in data[category['name']][i]):
                    logging.error('Double-check this item!')
                    error.append(asin)

                if verbose:
                    print(data[category['name']][i])
                    print()

                # delay
                time.sleep(random.random()*3+0.25)

            except Exception as e:
                data[category['name']][i] = [None]*14
                data[category['name']][i][2] = asin
                logging.error(e)
                error.append(asin)

    if verbose:
        print(json.dumps(data, default=str))

    return data, error

def write_to_excel(file_name, file_dir, data):
    if file_name is None:
        mode = 'w'
        version = datetime.now().strftime('%m%d%y_%H%M%S')
        file_name = os.path.join(version + '_output.xlsx')
    else:
        mode = 'a'

    file_path = os.path.join(file_dir, file_name)
    writer = pd.ExcelWriter(file_path, mode=mode, date_format='m/d/yy')

    headers_font = Font(underline='single', bold=True)
    hyperlink_font = Font(underline='single', color=colors.BLUE)

    for sheet in data:
        # get first empty row or make headers
        if sheet not in writer.book.sheetnames:
            # make headers
            sheet_headers = ['Date', 'Amazon Search Term', 'Product Title',
                             'URL', 'Quantity', 'Prime Or Not', 'Prime Price',
                             'Non-Prime Price', 'Non-Prime Shipping',
                             'Seller', 'Fulfilled By', 'Rating',
                             'Number Of Ratings', 'Packaging']
            writer.book.create_sheet(sheet)
            writer.book[sheet].append(sheet_headers)
            for cell in writer.book[sheet]['1'][:14]:
                cell.font = headers_font
            startrow = 1
        else:
            startrow = writer.book[sheet].max_row
            for cell in writer.book[sheet]['A']:
                if cell.value is None:
                    startrow = cell.row

        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

        # deserialize date if from JSON
        if data[sheet] and type(next(iter(data[sheet].values()))[0]) == str:
            for v in data[sheet].values():
                v[0] = datetime.strptime(v[0], '%Y-%M-%d').date()

        # write to sheet
        df = pd.DataFrame.from_dict(data[sheet], orient='index')
        df.to_excel(writer, sheet_name=sheet, startrow=startrow,
                    index_label=False, index=False, header=False)
        for cell in writer.book[sheet]['D'][startrow:]:
            cell.font = hyperlink_font

    writer.save()

if __name__ == '__main__':
    # argparse
    parser = argparse.ArgumentParser(description='Scrape! -Chris :)')

    parser.add_argument('-p', '--proxy', action='store_true', help='Use proxy rotation')
    parser.add_argument('-n', '--new',   action='store_true', help='Create new Excel file')
    parser.add_argument('-d', '--debug', help='ASIN of single product')
    parser.add_argument('-w', '--write', help='Write JSON data to file')

    args = parser.parse_args()

    # logging config
    logging.basicConfig(level=logging.INFO,
                        format='%(asctime)s.%(msecs)03d - %(levelname)s - %(message)s',
                        datefmt='%H:%M:%S')

    # set current working directory
    cwd = os.path.dirname(os.path.realpath(__file__))
    
    # load config
    logging.info('Loading config file')
    config_file = os.path.join(cwd, 'config.yml')
    with open(config_file, 'r') as f:
        config = yaml.safe_load(f)
    
    # set write location
    if args.new:
        data_file_name = None
    else:
        data_file_name = config['data_file']

    if args.write:
        logging.info('Writing JSON data...')
        write_to_excel(data_file_name, cwd, json.loads(args.write))
        logging.info('Write complete')
    else:
        # load useragents
        logging.info('Loading useragent file')
        with open(os.path.join(cwd, config['uagents']), 'r') as f:
            ua_pool = [line.strip() for line in f]

        # scrape
        logging.info('Beginning scrape...')
        if args.debug:
            data, error = scrape([{'name':'debug','asins':[args.debug]}],
                                  config['headers'], ua_pool, args.proxy,
                                  config['proxy_list'], 1, config['ip_check'],
                                  verbose=True)
        else:
            data, error = scrape(config['products'], config['headers'], ua_pool,
                                 args.proxy, config['proxy_list'],
                                 config['max_proxies'], config['ip_check'])

            # write to file
            try:
                write_to_excel(data_file_name, cwd, data)
                logging.info('Scrape complete')

                print('\n*** Double-check these items***:')
                for e in error:
                    print('- {}'.format(e))
            except Exception as e:
                logging.error(e)
                print('Data:')
                print(json.dumps(data))
                print('Errors:')
                print(error)
